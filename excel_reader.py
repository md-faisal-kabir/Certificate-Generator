from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def read_excel(stock,metal):
    # reading the cert template
    try:
        certbook = load_workbook(filename="Product Certification.xlsx")
        sheet_name = ''
        switcher = {'S': 'Cert Steel - Stock', 'N': 'Cert Nickel - Stock', 'A': 'Cert Aluminum - Stock', 'T': 'Cert Titanium - Stock' }
        if stock == '1':
            sheet_name = switcher[metal]
        else:
            sheet_name = 'Buyout Cert'
        cert_template = certbook[sheet_name]

        #reading the database
        db = load_workbook(filename="cert-db.xlsx")
        db_sheet = db.active
        prev_id = db_sheet['A'][len(db_sheet['A']) - 1].value
        return cert_template, db_sheet, certbook, db, prev_id
    except (PermissionError):
        print("*** CLOSE THE OPEN PRODUCTION CERTIFICATION.XLSX SHEET FIRST")

if __name__ == "__main__":
    while True:
        stock = input("( 1 ) STOCK OR ( 2 ) BUYOUT >>> ")
        check = ['1', '2']
        if stock in check:
            break
        else:
            print("Choose 1 or 2")
            continue

    while True:
        metal = input("METAL TYPE: (S)STEEL  (N)NICKEL (A)ALUMINUM (T)TITANIUM >>> ").upper()
        check = ['S', 'N', 'T', 'A']
        if metal in check:
            break
        else:
            print(" Invalid Entry. Try Again")
            continue
    
    cert_template, db_sheet, certbook, db, prev_id = read_excel(stock, metal)
    print(cert_template['O6'].value)
    print(db_sheet['A2'].value)
    print(prev_id)
